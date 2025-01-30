# from app import create_app
from flask import Flask, request, redirect, url_for, session,jsonify,render_template,make_response,flash,abort
from flask_cors import CORS
from flask_compress import Compress 
import requests
import json
from urllib.parse import urlencode
import pandas as pd 
import numpy as np
from datetime import datetime
from functools import wraps
from datetime import timedelta
import os 
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side,Font
import random
import string

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # 세션 관리를 위해 필요한 키 설정


CORS(app)
app.config['COMPRESS_LEVEL'] = 4 #압축레벨 설정 
Compress(app)

app.config['SAVE_FILES_DIR'] = '/static/data'
app.config['JSON_AS_ASCII'] = False 
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=60)



def nocache(view):
    @wraps(view)
    def no_cache(*args, **kwargs):
        response = make_response(view(*args, **kwargs))
        response.headers['Last-Modified'] = datetime.now()
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    return no_cache

# OAuth 2.0 인증 정보
client_id = '8zESjIVD54_8K5PA_dIu'
client_secret = 'wyWAH1uLEY'
redirect_uri = 'http://175.209.244.54:8080/callback'
auth_url = 'https://auth.worksmobile.com/oauth2/v2.0/authorize'
token_url = 'https://auth.worksmobile.com/oauth2/v2.0/token'
userinfo_url = 'https://apis.worksmobile.com/r/api/user/v1/userinfo'

def check_session_id(func):
    @wraps(func)
    def decorated_function(*args, **kwargs):
        if 'ID' not in session:  # 세션에 'id'가 없으면
            return render_template('backoffice-login.html')  # backoffice.html로 렌더링
        return func(*args, **kwargs)  # 세션에 'id'가 있으면 원래 함수 실행
    return decorated_function

def new_generate_random_key(df):
    """알파벳 3글자 + 숫자 4자리 난수 생성 (예: ABC1234)"""
    
    letters = ''.join(random.choices(string.ascii_uppercase, k=3))
    digits = ''.join(random.choices(string.digits, k=4))
    newkey = letters + digits

    if newkey in df['KEY'].unique():
        while newkey in df['KEY'].unique():
            print("겹치는거 있음?")
            letters = ''.join(random.choices(string.ascii_uppercase, k=3))
            digits = ''.join(random.choices(string.digits, k=4))
            newkey = letters + digits
        return newkey
    else:
        return newkey
    

@app.route('/', methods=['GET'])
@nocache
def idx():
    
    print("main gogo")

    return render_template("main.html")


@app.route('/untitled', methods=['GET'])
@nocache
def untitled():
    
    print("main gogo")

    return render_template("Untitled-2.html")

@app.route('/main', methods=['GET','POST'])
@nocache
def main():
    

    print("main gogo")
    return render_template("main.html")

@app.route('/employee', methods=['GET','POST'])
@nocache
def employee():
    
    print("employee gogo") 
    
    return render_template("main.html")

@app.route('/signin', methods=['GET','POST'])
@nocache
def signin():
    if(request.method=='GET'):
        file_path = './static/data/countries.json'
        with open(file_path, 'r') as f:
            countries = json.load(f)
            f.close()
        return render_template("signup.html", countries=countries['countries'])
    elif(request.method=='POST'):
        ##로그인 성공
        # ID랑 비번으로 계정 매핑  
        insadb = pd.read_excel("./static/data/INSA_DB.xlsx")
        insadb['PASSWORD'] = insadb['PASSWORD'].astype(str)

        file_path = './static/data/calendar_group.json'
        with open(file_path, 'r') as f:
            grouplist = json.load(f)
            f.close()

        login_data  = request.get_json()
        id = login_data['id']
        pw = login_data['pw']
        session['id'] = id
        session['pw'] = pw

        filterd_db = insadb[(insadb['ID']==id)&(insadb['PASSWORD']==pw)].reset_index(drop=True)

        if len(filterd_db)>0:
            site_group=[]
            for group in grouplist:
                if filterd_db['ID'][0] in group['members']:
                    site_group = np.append(site_group,group['groupName'])

            session['NAME'] = filterd_db['성명'][0]
            session['ID'] = filterd_db['ID'][0]
            session['ADMIN'] = filterd_db['ADMIN'][0]
            session['PLACE'] = site_group.tolist()
            name= session['NAME']
            admin = session['ADMIN']
            place = session['PLACE']
            place = place[0]
            print(id,name,admin,place)
            
            if session['ADMIN']=='MASTER':
                return render_template("infotab-master.html",id=id,name = name,admin=admin,place=session['PLACE'])
            elif session['ADMIN']=='ADMIN':
                work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")

                today = datetime.today().strftime("%Y-%m")
                work_info_filter = work_info[work_info['작업날짜'].str.contains(today).replace(np.NaN,False)]
                work_info_filter = work_info_filter[work_info_filter['작업장'] == place]
                work_info_json = work_info_filter[['작업날짜','작업장','위치_분류1','위치_분류2','작업_분류3','작업진행률']].to_json(orient='records', force_ascii=False)
            

                
                return render_template("infotab-admin.html",id=id,name = name,admin=admin,place=place,work_info_json=work_info_json)
            else:
                ## 나머지 데이터가 있는지 확인 
                db_cols = ['성명', '영문명', '체류비자', '국적', 'KEY_ID', '주소', '전화번호','자격증유무', '고혈압유무']
                null_data = filterd_db[db_cols].isna().sum(axis=1).values[0]
                if null_data!=0:
                    sign_data = 'False'
                ## 먼저 날짜 컬럼을 만들어야함 
                today_day = datetime.now().strftime(format = "%Y-%m-%d")
                today_time =  datetime.now().strftime(format = "%Y-%m-%d %H:%M")
                cols1 = today_day+"(출근)"
                cols2 = today_day+"(퇴근)"
                
                if cols1 in filterd_db.columns:
                    print("yess")
                    ## 여기에는 그냥 해당사람 시간 넣으면 됨 
                    Attend_t = filterd_db[cols1].values[0] 
                else:
                    Attend_t = 'none'
                
                if cols2 in filterd_db.columns:
                    print("yess")
                    ## 여기에는 그냥 해당사람 시간 넣으면 됨 
                    Leave_t = filterd_db[cols2].values[0]
                else:
                    Leave_t = 'none'
                print(Attend_t==np.nan)
                if pd.isna(Attend_t):
                    Attend_t='none'
                if pd.isna(Leave_t):
                    Leave_t='none'

                print('Attend,Leave',Attend_t,Leave_t)

                return jsonify(sign_data=sign_data)
        else:
            abort(400, description="Session ID not found")
        





@app.route('/infotab', methods=['GET','POST'])
@nocache
def infotab():
    
    print("infotab gogo") 
    if 'ADMIN' in session:
        
        ## 현재 작업 진행률 보여주기 
        if session['ADMIN']=='MASTER':
            name = session['NAME']
            id = session['ID']
            admin = 'MASTER'
            place = session['PLACE']
            ## 기존 정보들 다 가져감 
            
            return render_template("infotab-master.html",id=id,name = name,admin=admin,place=place)
        else:
            name = session['NAME']
            id = session['ID']
            admin = 'ADMIN'
            place = session['PLACE']
            work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")
            print(place)
            today = datetime.today().strftime("%Y-%m")
            work_info_filter = work_info[work_info['작업날짜'].str.contains(today).replace(np.NaN,False)]
            work_info_filter = work_info_filter[work_info_filter['작업장'] == place[0]]
            
            work_info_json = work_info_filter[['KEY','작업날짜','작업장','위치_분류1','위치_분류2','작업_분류2','작업_분류3','작업진행률','목표공수','공수']].to_json(orient='records', force_ascii=False)
        

            return render_template("infotab-admin.html",id=id,name = name,admin=admin,place=place,work_info_json=work_info_json)
    else:

        return render_template('backoffice-login.html')
    



@app.route('/gocalendar', methods=['GET'])
@nocache
@check_session_id
def gocalendar():
    insadb = pd.read_excel("./static/data/INSA_DB.xlsx")
    insadb['PASSWORD'] = insadb['PASSWORD'].astype(str)
    filterd_db = insadb[(insadb['ID']==session['id'])&(insadb['PASSWORD']==session['pw'])].reset_index(drop=True)
    ## 먼저 날짜 컬럼을 만들어야함 
    today_day = datetime.now().strftime(format = "%Y-%m-%d")
    today_time =  datetime.now().strftime(format = "%Y-%m-%d %H:%M")
    cols1 = today_day+"(출근)"
    cols2 = today_day+"(퇴근)"
    
    if cols1 in filterd_db.columns:
        print("yess")
        ## 여기에는 그냥 해당사람 시간 넣으면 됨 
        Attend_t = filterd_db[cols1].values[0] 
    else:
        Attend_t = 'none'
    
    if cols2 in filterd_db.columns:
        print("yess")
        ## 여기에는 그냥 해당사람 시간 넣으면 됨 
        Leave_t = filterd_db[cols2].values[0]
    else:
        Leave_t = 'none'
    print(Attend_t==np.nan)
    if pd.isna(Attend_t):
        Attend_t='none'
    if pd.isna(Leave_t):
        Leave_t='none'
    name = session['NAME']
    place = session['PLACE']
    if len(place)>=1:
        place= place[0]
    file_path = './static/data/Attend.json'
    with open(file_path, 'r') as f:
        attend = json.load(f)[place]['Attend']

    file_path = './static/data/Leave.json'
    with open(file_path, 'r') as f:
        leave = json.load(f)[place]['Leave']

    print('Attend,Leave',Attend_t,Leave_t)

    return render_template("calendar-employee.html",Attend_t=Attend_t,Leave_t=Leave_t,name = name,attend=attend,leave=leave)


@app.route('/survey-plan', methods=['GET','POST'])
@nocache
@check_session_id
def survey_plan():
    if(request.method=='GET'):
        
        
        name = session.get('NAME')
        if not name:
            return redirect(url_for('backoffice'))
        else:
            name = session['NAME']
            id = session['ID']
            insadb = pd.read_excel("./static/data/INSA_DB.xlsx")
            insadb_json = insadb[['성명','작업장']].to_json(orient='records', force_ascii=False)

            ## 기존 정보들 다 가져감 
            work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")
           
            file_path = './static/data/calendar_group.json'
            with open(file_path, 'r') as f:
                site_info = json.load(f)
            print('site_info',site_info)
            for key in range(len(site_info)):
                if id in site_info[key]['members']:
                    site = site_info[key]['groupName']
                else:
                    site=''
            work_info['작업날짜'] = work_info['작업날짜'].astype(str)
            work_info_filter = work_info.loc[(work_info['작업날짜']>='2025-01-01')&(work_info['작업장']==site)&(work_info['작업진행률']!='100%'),['도급/직영','위치_분류1','위치_분류2','작업_분류1','작업_분류2','작업_분류3','작업날짜','작업진행률','작업장']]
            print(work_info_filter)
            work_info_filter['작업상세'] = work_info_filter['도급/직영']+"_"+work_info_filter['위치_분류1']+"층_"+work_info_filter['위치_분류2']+"_"+work_info_filter['작업_분류3']
            work_info_json = work_info_filter[['작업날짜','작업장','작업상세','작업진행률']].to_json(orient='records', force_ascii=False)
            


            return render_template("survey-plan.html",name =name,insadb= insadb_json,work_info_json=work_info_json)    
    elif(request.method=='POST'):

        
        survey_data  = request.get_json()
        taskresult = survey_data['taskresult'].split("\n")
        date = survey_data['date'].split(" ")[0]
        username = survey_data['username']

        work_info = pd.read_excel("./static/data/작업공수/작업계획.xlsx")
 
        file_path = "./static/data/test.json"
        with open(file_path, 'r',encoding='utf-8') as outfile:
            old_df = pd.DataFrame(json.loads(json.load(outfile)))
        
        for task in taskresult:
            task_info = task.split("_")
            if len(task)>1:
                temp_task = pd.DataFrame({"작업날짜":[date],
                                          '작업장':[task_info[0]],
                                        "위치_분류1":[task_info[2]],
                                        "위치_분류2":[task_info[3]],
                                        "작업_분류2":[task_info[4]], ## 작업분류1 알폼/코아 필요한데 
                                        "작업_분류3":[task_info[5]],
                                        "목표공수":[float(task_info[6])],
                                        "도급/직영":[task_info[1]]
                                        })
                key_column = ['작업장', '위치_분류1', '위치_분류2', '작업_분류2', '도급/직영', '작업_분류3']
                dup = work_info.loc[(work_info[key_column]==temp_task[key_column].values).sum(axis=1)==6]
   
                if len(dup)!=0:
                    print('error')
                else:
                    new_key = new_generate_random_key(work_info)
                    temp_task['KEY'] = new_key
                    work_info = pd.concat([work_info,temp_task]).reset_index(drop=True)

                    new_val = pd.DataFrame({"구간":[task_info[3]],"층":[task_info[2]],"세부구간":[task_info[4]],"중분류":[task_info[1]],"작업내용":[task_info[5]],"공수":[float(task_info[6])],"작업진행률":[task_info[7]],"작업날짜":[date],"직종":["??"],"작업장명":[task_info[0]]})
                    new_df = pd.concat([old_df,new_val]).reset_index(drop=True)
        

        with open(file_path, 'w',encoding='utf-8') as outfile:
            json.dump(new_df.to_json(orient='records'), outfile, ensure_ascii=False, indent=4)

        if(len(work_info)>1):
            print('worker_info',work_info)
            work_info.to_excel("./static/data/작업공수/작업계획.xlsx", index=False)
        return jsonify("success")


@app.route('/survey', methods=['GET','POST'])
@nocache
@check_session_id
def survey():
    if(request.method=='GET'):
        
        
        name = session.get('NAME')
        if not name:
            return redirect(url_for('backoffice'))
        else:
            name = session['NAME']
            id = session['ID']
            insadb = pd.read_excel("./static/data/INSA_DB.xlsx")
            insadb_json = insadb[['성명','작업장']].to_json(orient='records', force_ascii=False)

            ## 기존 정보들 다 가져감 
            work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")
           
            # file_path = './static/data/calendar_group.json'
            # with open(file_path, 'r') as f:
            #     site_info = json.load(f)
                
            # for key in range(len(site_info)):
            #     if id in site_info[key]['members']:
            #         site = site_info[key]['groupName']
            #     else:
            #         site='김포'
            site = session['PLACE'][0]
            print(site)
            work_info['작업날짜'] = work_info['작업날짜'].astype(str)
            print(work_info)
            work_info_filter = work_info.loc[(work_info['작업날짜']>='2025-01-01')&(work_info['작업장']==site)&(work_info['작업진행률']!='100%'),['도급/직영','위치_분류1','위치_분류2','작업_분류1','작업_분류2','작업_분류3','작업날짜','작업진행률','작업장','목표공수','공수','KEY']]
            print(work_info_filter)
            work_info_filter['작업상세'] = work_info_filter['도급/직영']+"_"+work_info_filter['위치_분류1']+"층_"+work_info_filter['위치_분류2']+"_"+work_info_filter['작업_분류3']
            work_info_json = work_info_filter[['작업날짜','작업장','작업상세','작업진행률','KEY','목표공수','공수']].to_json(orient='records', force_ascii=False)
            


            return render_template("survey.html",name =name,insadb= insadb_json,work_info_json=work_info_json)    
    elif(request.method=='POST'):

        def save_double_column_df(df, file_name, startrow = 0, **kwargs):
            '''Function to save doublecolumn DataFrame, to xlwriter'''
            # https://stackoverflow.com/questions/52497554/blank-line-below-headers-created-when-using-multiindex-and-to-excel-in-python
            
            # inputs:
            # df - pandas dataframe to save
            # xl_writer - book for saving
            # startrow - row from wich data frame will begins
            # **kwargs - arguments of `to_excel` function of DataFrame`
            with pd.ExcelWriter(file_name, mode='w') as writer:
                df.drop(df.index).to_excel(writer, startrow = startrow, **kwargs)
                df.to_excel(writer, startrow = startrow + 1, header = False, **kwargs)
        ## 작업
        survey_data  = request.get_json()
        # workersinfo = survey_data['workersinfo']
        taskresult = survey_data['taskresult'].split("\n")
        # additional = survey_data['additional'].split("\n")
        date = survey_data['date'].split(" ")[0]
        username = survey_data['username']
        # print('workersinfo',workersinfo)

        # workersinfo_lines = workersinfo.strip().split("\n")

        # 각 줄을 파싱하여 딕셔너리로 변환
        # parsed_data = []
        # for line in workersinfo_lines:
        #     parts = line.split('|')  # '|'를 기준으로 데이터 분리
        #     parsed_data.append({
        #         '이름': parts[0].split(': ')[1].strip(),  # '행 1: ' 제거 후 이름 추출
        #         '공수': parts[1].strip(),                  # ID 추출
        #         '작업진행률': parts[2].strip(),     # 완료율 추출
        #         '작업': parts[3].strip(),    # 작업 설명 추출
        #         'Action': parts[4].strip()               # 작업 동작 추출
        #     })
        # df_workersinfo = pd.DataFrame(parsed_data)

        work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")
        # worker_info  = pd.read_excel("./static/data/작업공수/작업자입력.xlsx",header=[0,1])
 

        # if("Unnamed" in worker_info.columns[0][1]):
        #     worker_info = worker_info.iloc[:,1:]

        # new_columns = [
        #     (("" if "Unnamed" in str(outer) else outer.strftime('%Y-%m-%d') if isinstance(outer, datetime) else outer), inner)
        #     for outer, inner in worker_info.columns
        # ]

        # # 새로운 컬럼 이름을 적용
        # worker_info.columns = pd.MultiIndex.from_tuples(new_columns)

        

        # if (date, '측정') in worker_info.columns:
        #     print("있으니깐 패스")
        # else:
        #     data2 = pd.DataFrame({
        #         (date, '측정'): [],
        #         (date, '실'): [],
        #     })
        #     worker_info = pd.concat([worker_info,data2],axis=0)

        # ## 사람만큼 공수 추가 (일반은 1, 오전,특수는 0.5)
        # for worker_name in df_workersinfo['이름']:
        #     if worker_info.loc[worker_info[('','성명')]==worker_name,(date, '측정')].isna().any():
        #         worker_info.loc[worker_info[('','성명')]==worker_name,(date, '측정')] = 0
        #     worker_info.loc[worker_info[('','성명')]==worker_name,(date, '측정')] += float(df_workersinfo.loc[df_workersinfo['이름']==worker_name,'공수'])

        file_path = "./static/data/test.json"
        with open(file_path, 'r',encoding='utf-8') as outfile:
            old_df = pd.DataFrame(json.loads(json.load(outfile)))
        
        for task in taskresult:
            task_info = task.split("_")
            if len(task)>1:
                temp_task = pd.DataFrame({"작업날짜":[date],
                                          '작업장':[task_info[0]],
                                        "위치_분류1":[task_info[2]],
                                        "위치_분류2":[task_info[3]],
                                        "작업_분류2":[task_info[4]],
                                        "작업_분류3":[task_info[5]],
                                        "공수":[float(task_info[6])],
                                        "도급/직영":[task_info[1]],
                                        "작업진행률":[task_info[7]]
                                        })
                
                key_column = ['작업장', '위치_분류1', '위치_분류2', '작업_분류2', '도급/직영', '작업_분류3']
                dup = work_info.loc[(work_info[key_column]==temp_task[key_column].values).sum(axis=1)==6]
   
                if len(dup)!=0:
                    print('error')
                else:
                    new_key = new_generate_random_key(work_info)
                    temp_task['KEY'] = new_key
                    work_info = pd.concat([work_info,temp_task]).reset_index(drop=True)

                # work_info = pd.concat([work_info,temp_task]).reset_index(drop=True)
                new_val = pd.DataFrame({"구간":[task_info[3]],"층":[task_info[2]],"세부구간":[task_info[4]],"중분류":[task_info[1]],"작업내용":[task_info[5]],"공수":[float(task_info[6])],"작업진행률":[task_info[7]],"작업날짜":[date],"직종":["??"],"작업장명":[task_info[0]]})
                new_df = pd.concat([old_df,new_val]).reset_index(drop=True)
        #worker_info.columns = pd.MultiIndex.from_tuples(worker_info.columns)
        

        with open(file_path, 'w',encoding='utf-8') as outfile:
            json.dump(new_df.to_json(orient='records'), outfile, ensure_ascii=False, indent=4)

        # if(len(worker_info)>1):
        #     print('worker_info',worker_info)
        #     #worker_info.to_excel()
        #     save_double_column_df(worker_info,"./static/data/작업공수/작업자입력.xlsx")
        if(len(work_info)>1):
            print('worker_info',work_info)
            work_info.to_excel("./static/data/작업공수/작업입력.xlsx", index=False)
        return jsonify("success")



@app.route('/survey_start', methods=['POST'])
@nocache
def survey_start():
    start_date = request.get_json()
    print(start_date)
    session['CALENDAR_DATE'] = start_date
    return jsonify({"success": True}), 200


@app.route('/calendar', methods=['GET'])
@nocache
def calendar():
    print(session)
    if 'ADMIN' in session:
        print("session['ADMIN']",session['ADMIN'])
        if session['ADMIN']=='MASTER':
            name = session['NAME']
            id = session['ID']
            admin = 'MASTER'
            place = session['PLACE']
            if len(place)>=1:
               place= place[0]
            file_path = './static/data/Attend.json'
            with open(file_path, 'r') as f:
                attend = json.load(f)[place]['Attend']

            file_path = './static/data/Leave.json'
            with open(file_path, 'r') as f:
                leave = json.load(f)[place]['Leave']
            
            return render_template('calendar-master.html',id=id,name = name,admin=admin,attend=attend,leave=leave,place = session['PLACE'])
        elif session['ADMIN']=='ADMIN':
            name = session['NAME']
            admin = 'ADMIN'
            id = session['ID']
            place = session['PLACE']
            print(place)
            if len(place)>=1:
               place= place[0]
            print(place)
            file_path = './static/data/Attend.json'
            with open(file_path, 'r') as f:
                attend = json.load(f)[place]['Attend']

            file_path = './static/data/Leave.json'
            with open(file_path, 'r') as f:
                leave = json.load(f)[place]['Leave']
            
            return render_template('calendar-admin.html',id=id,name = name,admin=admin,attend=attend,leave=leave,place = session['PLACE'])

        else:
            name = session['NAME']
            place = session['PLACE']
            if len(place)>=1:
               place= place[0]
            admin='N'
            file_path = './static/data/Attend.json'
            with open(file_path, 'r') as f:
                attend = json.load(f)[place]['Attend']

            file_path = './static/data/Leave.json'
            with open(file_path, 'r') as f:
                leave = json.load(f)[place]['Leave']

            insadb = pd.read_excel("./static/data/INSA_DB.xlsx")

    
        
            filterd_db = insadb[insadb['ID']==session['ID']].reset_index(drop=True)
            ## 먼저 날짜 컬럼을 만들어야함 
            today_day = datetime.now().strftime(format = "%Y-%m-%d")
            today_time =  datetime.now().strftime(format = "%Y-%m-%d %H:%M")
            cols1 = today_day+"(출근)"
            cols2 = today_day+"(퇴근)"
            print(filterd_db)
            if cols1 in filterd_db.columns:
                print("yess")
                ## 여기에는 그냥 해당사람 시간 넣으면 됨 
                Attend_t = filterd_db[cols1].values[0]
            else:
                Attend_t = 'none'
            
            if cols2 in filterd_db.columns:
                print("yess")
                ## 여기에는 그냥 해당사람 시간 넣으면 됨 
                Leave_t = filterd_db[cols2].values[0]
            else:
                Leave_t = 'none'
         

            if pd.isna(Attend_t):
                Attend_t='none'
            if pd.isna(Leave_t):
                Leave_t='none'
                
            print('Attend,Leave',Attend_t,Leave_t)
            print('attend',attend,leave)
            return render_template('calendar-employee.html',name = name,admin=admin,attend=attend,leave=leave,Attend_t=Attend_t,Leave_t=Leave_t)
        
    else:
        return render_template('backoffice-login.html')


@app.route('/calendar_save', methods=['POST'])
@nocache
def calendar_save():
    print(request.get_json())
    data = request.get_json()['data']
    file_path = './static/data/calendar.json'
    data = json.loads(data)
    with open(file_path, 'w') as outfile:
      json.dump(data, outfile, indent=4)
    
    return jsonify({"success": True}), 200


@app.route('/calendar_remove', methods=['POST'])
@nocache
def calendar_remove():
    print(request.get_json())
    data = request.get_json()['data']
    file_path = './static/data/calendar.json'
    data = json.loads(data)
    

    with open(file_path, 'w') as outfile:
      json.dump(data, outfile, indent=4)
    
    return jsonify({"success": True}), 200

@app.route('/calendar_edit', methods=['POST'])
@nocache
def calendar_edit():
    print(request.get_json())
    print('request',request.form)
    data = request.get_json()['data']
    file_path = './static/data/calendar.json'
    data = json.loads(data)
    with open(file_path, 'w') as outfile:
      json.dump(data, outfile, indent=4)
    
    return jsonify({"success": True}), 200


@app.route('/save_calendar_group', methods=['POST'])
@nocache
def save_calendar_group():
    print(request.get_json())
    print('request',request.form)
    data = request.get_json()
    file_path = './static/data/calendar_group.json'
    # data = json.loads(data)
    print(data)
    with open(file_path, 'w') as outfile:
      json.dump(data, outfile, indent=4)
    
    return jsonify({"success": True}), 200


@app.route('/productivity', methods=['GET','POST'])
@nocache
def productivity():
    if(request.method=='GET'):
        print("main gogo")
        work_info = pd.read_excel("./static/data/작업공수/작업입력.xlsx")

        print(work_info[['작업날짜','작업장','위치_분류1','위치_분류2','작업_분류3','작업진행률']] )
       
        work_info_json = work_info[['작업날짜','작업장','위치_분류1','위치_분류2','작업_분류3','작업진행률']].to_json(orient='records', force_ascii=False)
        
        ## 작업 생산성 데이터 
        anyang = pd.read_excel("D:/네이버웍스/NAVER WORKS Drive/. Public_Root/외주@2001000000403897/2024년 11월 안양 태일 제출서류(1129까지).xlsm",sheet_name='수익 분석')
        songdo = pd.read_excel("D:/네이버웍스/NAVER WORKS Drive/. Public_Root/외주@2001000000403897/2024년 11월 송도 디아이이엔씨 제출서류(1129까지).xlsm",sheet_name='수익 분석')
        magock = pd.read_excel("D:/네이버웍스/NAVER WORKS Drive/. Public_Root/외주@2001000000403897/2024년 11월 마곡 신세계 제출서류(1129까지).xlsm",sheet_name='수익 분석')

        anyang.columns = np.append(['TYPE'],anyang.columns[1:])
        songdo.columns=np.append(['TYPE'],songdo.columns[1:])
        magock.columns=np.append(['TYPE'],magock.columns[1:])
        anyang['TYPE'] = anyang['TYPE'].str.replace(" ","")
        songdo['TYPE'] = songdo['TYPE'].str.replace(" ","")
        magock['TYPE'] = magock['TYPE'].str.replace(" ","")

        anyang['수익률'] = round((anyang['수익']/anyang['금액'])*100,2)
        songdo['수익률'] = round((songdo['수익']/songdo['금액'])*100,2)
        magock['수익률'] = round((magock['수익']/magock['금액'])*100,2)
        anyang['생산성'] = anyang['생산성'].round(2)
        songdo['생산성'] = songdo['생산성'].round(2)
        magock['생산성'] = magock['생산성'].round(2)
        

        temp1 = anyang.loc[anyang.TYPE.isin(['총공사','도급','직영']),['TYPE','생산성','수익률']].set_index('TYPE')
        temp2 = songdo.loc[songdo.TYPE.isin(['총공사','도급','직영']),['TYPE','생산성','수익률']].set_index('TYPE')
        temp3 = magock.loc[magock.TYPE.isin(['총공사','도급','직영']),['TYPE','생산성','수익률']].set_index('TYPE')
        
        anyang_data = temp1.to_dict(orient='records')
        songdo_data = temp2.to_dict(orient='records')
        magock_data = temp3.to_dict(orient='records')

        # 2) 하나의 파이썬 dict로 합치기
        combined_dict =pd.DataFrame({
            '안양': anyang_data,
            '송도': songdo_data,
            '마곡': magock_data
        },index=temp1.index)

        ## 더미 데이터 만들기 
        combined_dict2 =pd.DataFrame({
            '김민수': anyang_data,
            '장재룡': songdo_data,
            'Aslera': magock_data
        },index=temp1.index)
        


        product_json = combined_dict.to_json( force_ascii=False)
        personal_json = combined_dict2.to_json( force_ascii=False)
        
        return render_template("productivity.html",work_info_json=work_info_json,product_json=product_json,personal_json=personal_json)

##record-employee 
@app.route('/record-employee', methods=['GET',"POST"])
@nocache
def record_employee():
    if(request.method=='POST'):
        id = request.get_json()['id']
        print(id)
        file_path = './static/data/review.json'

        with open(file_path, 'r') as infile:
            data = json.load(infile)
      
        print(data)
        review = data[id]
        return jsonify(review =review)    
    elif(request.method=='GET'):
        insa = pd.read_excel("D:/네이버웍스/NAVER WORKS Drive/. Public_Root/외주@2001000000403897/20241130 인사기록카드.xlsx",sheet_name='개인DB(수정금지)')
        insa.loc[(~insa['기초안전보건교육 이수증'].isna())&(~insa['건설업 취업인정증'].isna())&(~insa['국가기술자격증 종류'].isna()),'등급'] = '고급'
        insa.loc[(~insa['기초안전보건교육 이수증'].isna())&(~insa['건설업 취업인정증'].isna())&(insa['등급'].isna()),'등급'] = '중급'
        insa.loc[insa['등급'].isna(),'등급'] = '초급'


        ## 첫 정보 
        info1 = insa[['고유번호','성명','국적','등급']].set_index('고유번호').transpose().to_json(force_ascii=False)

        ## 상세보기 경력

        ##상세보기 자격증
        temp1 = insa[['기초안전보건교육 이수증 발급날짜', '기초안전보건교육 이수증 발급기관', '기초안전보건교육 이수증']].to_dict(orient='records')
        temp2 = insa[['건설업 취업인정증 발급날짜', '건설업 취업인정증 유효기간', '건설업 취업인정증 발급기관', '건설업 취업인정증']].to_dict(orient='records')
        temp3 = insa[['국가기술자격증 발급날짜', '국가기술자격증 발급기관','국가기술자격증 종류']].to_dict(orient='records')
        temp4 = insa[['기능습득 교육 이수증 발급기관', '기능습득 교육 이수증', '기능습득 교육 종류']].to_dict(orient='records')
        temp5 = insa[['3톤미만지게차 건설기계조종사면허증 발급날짜', '3톤미만지게차 건설기계조종사면허증 발급기관','3톤미만지게차 건설기계조종사면허증']].to_dict(orient='records')


        df_cert = pd.DataFrame({
            "기초안전":temp1,
            "건설업":temp2,
            "국가기술":temp3,
            "기능습득":temp4,
            "지게차":temp5
        },index=insa['고유번호'])

        info2 = df_cert.transpose().to_json( force_ascii=False)
    
        print("employee gogo")
        return render_template("record-employee.html",info1=info1,info2=info2)

@app.route('/record-save', methods=["POST"])
@nocache
def record_save():
    saved_data = request.get_json()
    try:
        ID = session['id']
    except:
        ID = 'ADMIN9'
    file_path = './static/data/review.json'
    with open(file_path, 'r') as infile:
        data = json.load(infile)
    userid = saved_data['user']
    data[userid]['ID'].append(ID)
    data[userid]['Comment'].append(saved_data['Comment'])
    data[userid]['rate'].append(saved_data['rate'])
    with open(file_path, 'w') as outfile:
      json.dump(data, outfile, indent=4)

    return jsonify('success')    



@app.route('/create_event',methods=['POST'])
@nocache
def create_event():
    access_token = session.get('access_token')
    if not access_token:
        return redirect(url_for('main'))

    print(access_token)

    user_id = 'me'
    temp = requests.get(f"https://www.worksapis.com/v1.0/users/{user_id}/calendar",
                 headers={
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    })

    #calendar_id = temp.json()['calendarId']
    calendar_id = 'c_300029314_8d914568-af3f-4ac8-b9eb-e968f8547c08'
    event_url = f'https://www.worksapis.com/v1.0/users/{user_id}/calendars/{calendar_id}/events'
    print("event_url",event_url)
    event = json.loads(request.get_data())
    event['start']={'dateTime':'2024-07-06T14:00:00', 'timeZone':'Asia/Seoul'}
    event['description'] = event['description'].replace(":",":\n\n ").replace(",","\n ").replace('"',"").replace("workplace","작업현장").replace("numworkers","작업인원").\
        replace("workersnormal","일반 근무자").replace("workershard","추가 근무자").replace("workersmorning","오전근무자").\
        replace("taskresult","작업내용").replace("specialnotes","특이사항").replace("overtime","추가근무")

    data =json.dumps({"eventComponents":[event]})

    response = requests.post(event_url, headers={
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    },
    data=data)


    calendar_url = f'https://www.worksapis.com/v1.0/users/{user_id}/calendars/{calendar_id}/events'



    # 캘린더 이벤트 요청
    cal_list = requests.get(calendar_url, headers={
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    },params={"fromDateTime":"2024-07-01T10:00:00%2B09:00",
                                                                'untilDateTime':"2024-07-06T10:00:00%2B09:00"})
    print('cal_listcal_listcal_list',cal_list.json())


    if response.status_code == 200:
        return 'Event created successfully'
    else:
        print(response.text)
        return f"Failed to create event: {response.text}"

@app.route('/backoffice', methods=['GET'])
@nocache
def backoffice():
    if 'ID' in session:
        print(session['ID'])
        return redirect("/infotab")
    else:
      return render_template("backoffice-login.html")

@app.route('/mainback', methods=['GET'])
@nocache
def mainback():
    return render_template("backoffice-main.html")


@app.route('/check_admin', methods=['GET'])
@nocache
def check_admin():
    admin = session['ADMIN']
    place = session['PLACE']
    print('admin',admin)
    print('place',place)
    return jsonify({"ADMIN":admin,"PLACE":place})

@app.route('/attend_qr', methods=['GET'])
@nocache
def attendqr():

    return render_template('attend_qr.html')

@app.route('/attend-abled', methods=['POST'])
@nocache
def attendabled():
    print(request.get_json())
    data = request.get_json()
    file_path = './static/data/Attend.json'
    place = session['PLACE']
    if len(place)>=1:
        place= place[0]

    with open(file_path, 'r') as infile:
        df_attend = json.load(infile)
    df_attend[place] = data

    with open(file_path, 'w') as outfile:
      json.dump(df_attend, outfile, indent=4)
    return jsonify('success')

@app.route('/attend-disabled', methods=['POST'])
@nocache
def attenddisabled():
    print(request.get_json())
    data = request.get_json()
    file_path = './static/data/Attend.json'
    place = session['PLACE']
    if len(place)>=1:
        place= place[0]


    with open(file_path, 'r') as infile:
        df_attend = json.load(infile)
    df_attend[place] = data

    with open(file_path, 'w') as outfile:
      json.dump(df_attend, outfile, indent=4)
    return jsonify('success')


@app.route('/leave-abled', methods=['POST'])
@nocache
def leaveabled():
    print(request.get_json())
    data = request.get_json()
    file_path = './static/data/Leave.json'

    place = session['PLACE']
    if len(place)>=1:
        place= place[0]

    with open(file_path, 'r') as infile:
        df_leave = json.load(infile)
    df_leave[place] = data

    with open(file_path, 'w') as outfile:
      json.dump(df_leave, outfile, indent=4)
    return jsonify('success')

@app.route('/leave-disabled', methods=['POST'])
@nocache
def leavedisabled():
    print(request.get_json())
    data = request.get_json()
    file_path = './static/data/Leave.json'

    place = session['PLACE']
    if len(place)>=1:
        place= place[0]

    with open(file_path, 'r') as infile:
        df_leave = json.load(infile)
    df_leave[place] = data

    with open(file_path, 'w') as outfile:
      json.dump(df_leave, outfile, indent=4)

    return jsonify('success')


@app.route('/attend-check', methods=['POST'])
@nocache
def attendheck():
    
    # ID랑 비번으로 계정 매핑  
    insadb = pd.read_excel("./static/data/INSA_DB.xlsx")

    ## 먼저 날짜 컬럼을 만들어야함 
    today_day = datetime.now().strftime(format = "%Y-%m-%d")
    today_time =  datetime.now().strftime(format = "%Y-%m-%d %H:%M")
    cols = today_day+"(출근)"
    if cols in insadb.columns:
        print("yess")
        ## 여기에는 그냥 해당사람 시간 넣으면 됨 
        insadb.loc[insadb.ID==session['ID'],cols] = today_time

    else:
        print('no')
        insadb[cols]=""
        insadb.loc[insadb.ID==session['ID'],cols] = today_time

    insadb.to_excel("./static/data/INSA_DB.xlsx",index=False)



    return jsonify('success')




@app.route('/leave-check', methods=['POST'])
@nocache
def leaveheck():
    
    # ID랑 비번으로 계정 매핑  
    insadb = pd.read_excel("./static/data/INSA_DB.xlsx")

    ## 먼저 날짜 컬럼을 만들어야함 
    today_day = datetime.now().strftime(format = "%Y-%m-%d")
    today_time =  datetime.now().strftime(format = "%Y-%m-%d %H:%M")
    cols = today_day+"(퇴근)"
    if cols in insadb.columns:
        print("yess")
        ## 여기에는 그냥 해당사람 시간 넣으면 됨 
        insadb.loc[insadb.ID==session['ID'],cols] = today_time

    else:
        print('no')
        insadb[cols]=""
        insadb.loc[insadb.ID==session['ID'],cols] = today_time

    insadb.to_excel("./static/data/INSA_DB.xlsx",index=False)

    return jsonify('success')

@app.route('/personal_agree', methods=['GET'])
@nocache
def personal_agree():
    return render_template("personal_agree.html")

@app.route('/signup', methods=['GET'])
@check_session_id
@nocache
def signup():
    
    file_path = './static/data/countries.json'
    with open(file_path, 'r') as f:
        countries = json.load(f)
        f.close()

    # ID랑 비번으로 계정 매핑  
    insadb = pd.read_excel("./static/data/INSA_DB.xlsx")
    insadb['PASSWORD'] = insadb['PASSWORD'].astype(str)
    if 'id' in session and 'pw' in session:
        id = str(session['id']) 
        pw = str(session['pw'])
        print('id',id,pw)
        print(insadb)
        print(insadb[insadb['ID']==id])
        try:
            filterd_db = insadb[(insadb['ID']==id)&(insadb['PASSWORD']==pw)].reset_index(drop=True)
        except:
            filterd_db = insadb[(insadb['ID']==id)&(insadb['PASSWORD']==pw)].reset_index(drop=True)
        print(filterd_db)
        if len(filterd_db)>0:## 정보가 있는경우 
            filterd_db= filterd_db.fillna("")
            user_name = filterd_db['성명'].values[0]
            user_contact = filterd_db['전화번호'].values[0]
            user_nationality = filterd_db['국적'].values[0]
            user_credential = filterd_db['KEY_ID'].values[0]
            user_visa = filterd_db['체류비자'].values[0]
            user_address = filterd_db['주소'].values[0]
            user_certificate = filterd_db['자격증유무'].values[0]
            user_highBlood = filterd_db['고혈압유무'].values[0]
            info_list = {'user_id':id,
                        'user_name':user_name,
                        'user_contact':user_contact,
                        'user_nationality':user_nationality,
                        'user_credential':user_credential,
                        'user_visa':user_visa,
                        'user_address':user_address,
                        'user_certificate':user_certificate,
                        'user_highBlood':user_highBlood}
        else: 
            info_list={}

        return render_template("signup.html", countries=countries['countries'],info_list=info_list)
    else:

        info_list={}
        return render_template("signup.html", countries=countries['countries'],info_list=info_list)

@app.route('/logout', methods=['POST'])
@nocache
def logout():
    session.clear()
    print("맞아?")
    return jsonify("success")
    


@app.route('/register', methods=['POST'])
@nocache
def register():
    # ID랑 비번으로 계정 매핑  
    insadb = pd.read_excel("./static/data/INSA_DB.xlsx")

    register_data  = request.get_json()
    user_id = register_data['user_id']
    user_pw = register_data['user_pw']
    user_name = register_data['user_name']
    user_contact = register_data['user_contact']
    user_nationality = register_data['user_nationality']
    user_credential = register_data['user_credential']
    user_visa = register_data['user_visa']
    user_address = register_data['user_address']
    user_certificate = register_data['user_certificate']
    user_highBlood = register_data['user_highBlood']
    radio_nationality = register_data['radio_nationality']

    if radio_nationality=='총괄':
        registerd_data = pd.DataFrame({
            "성명":[user_name],
            "영문명":[user_name],
            "체류비자":[user_visa],
            "국적":[user_nationality],
            "KEY_ID":[user_credential],
            "주소":[user_address],
            "전화번호":[user_contact],
            "자격증유무":[user_certificate],
            "고혈압유무":[user_highBlood],
            "ID":[user_id],
            "PASSWORD":[user_pw],
            "ADMIN":["ADMIN"]
            
        })
    else:
        registerd_data = pd.DataFrame({
            "성명":[user_name],
            "영문명":[user_name],
            "체류비자":[user_visa],
            "국적":[user_nationality],
            "KEY_ID":[user_credential],
            "주소":[user_address],
            "전화번호":[user_contact],
            "자격증유무":[user_certificate],
            "고혈압유무":[user_highBlood],
            "ID":[user_id],
            "PASSWORD":[user_pw],
            "ADMIN":["N"]
            
        })

    if user_id in insadb['ID'].unique():
        # insadb = insadb[insadb.ID!=user_id]
        # insadb_new = pd.concat([insadb,registerd_data]).reset_index(drop=True)
        insadb_new = insadb.copy()
        insadb.loc[insadb.ID==user_id,
                   ["성명","영문명","체류비자","국적","KEY_ID",
                    "주소","전화번호","자격증유무","고혈압유무","ID","PASSWORD"]] = [user_name,user_name,user_visa,user_nationality,user_credential,user_address,user_contact,user_certificate,user_highBlood,user_id,user_pw]
        insadb.to_excel("./static/data/INSA_DB.xlsx",index=False)
    else:
        insadb_new = pd.concat([insadb,registerd_data]).reset_index(drop=True)
        insadb_new.to_excel("./static/data/INSA_DB.xlsx",index=False)

    print(insadb_new)
    
    return jsonify("success")


@app.route('/upload-image', methods=['POST'])
@check_session_id
def upload_image():
    if 'image' not in request.files:
        return jsonify({'success': False, 'message': '이미지가 없습니다.'})
    id = session['ID']
    print( request.files)
    text_data = request.form.get('textData') 
    print(text_data)
    file = request.files['image']
    
    print('request.files',request.files)
    print('img file',file)
    # print(file.content)
    print(id)
    type = file.filename.split(".")[1]
    file.filename = id + "_"+text_data +"." + type
    
    if file.filename == '':
        return jsonify({'success': False, 'message': '파일 이름이 비어 있습니다.'})
    
    path = './static/data/개인화폴더/'+str(id)
    # 파일 저장
    file_path = os.path.join(path, file.filename)
    file.save(file_path)
    try:
        
        return jsonify({'success': True, 'message': '이미지가 성공적으로 저장되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})



@app.route('/work_sheet', methods=['POST'])
@check_session_id
def work_sheet():
    id = session['ID']
    date = request.get_json()['date']
    print(date)
   # JSON 파일을 열고 데이터 읽기
    with open('./static/data/calendar.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    temp2=[]
    for i in range(len(data)):
        temp = data[i]
        
        if (pd.to_datetime(temp['start'])>=pd.to_datetime(date))&(pd.to_datetime(temp['start'])<pd.to_datetime(date) + relativedelta(months=1)):
            print(temp)
            temp2 = np.append(temp2,[temp])
    print(temp2)
    for i in range(len(temp2)):
        temp2[i]['description2']= ""
        for tmp in temp2[i]['description'].split("\n작업장"):
            print("tmp??",tmp)
            if "undefined" not in tmp:
                if temp2[i]['description2']!="":
                    temp2[i]['description2'] = temp2[i]['description2']+"\n"+tmp
                else: 
                    temp2[i]['description2'] = tmp

    print("temp2",temp2)
    # temp2[2]['description2'] = temp2[2]['description2'].replace("\n연장근무","")
    # temp2[2]['description2'] = temp2[2]['description2'][:-3]
    # temp2[3]['description2'] = temp2[3]['description2'].replace("\n연장근무","")
    # temp2[3]['description2'] = temp2[3]['description2'][:-3]
    iter_sheet = [50 * i for i in range(0,30)]
    wb = Workbook()
    ws = wb.active
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="D9D9D9",
                        end_color="D9D9D9",
                        fill_type="solid")


    ws.title = "작업일지"


    for iter in range(len(temp2)):
        print('iter',iter)
        today_worker = []
        cell_row = iter_sheet[iter]
        ws.merge_cells("A"+str(cell_row+1)+":S"+str(cell_row+1))  # A1 ~ F1까지 병합
        ws["A"+str(cell_row+1)] = "작 업 일 지"
        ws["A"+str(cell_row+1)].alignment = Alignment(horizontal='center')
        ws["A"+str(cell_row+1)].font = Font(size=14, bold=True)

        temp_parse = temp2[iter]['description2'].split("작업장")[1:]
        parse_i = temp_parse[0].split("\n")
        print('parse_i',parse_i)
        site= parse_i[0].split(":")[1].replace(" ","") ##작업장
        start = temp2[iter]['start']

        ws.merge_cells("A"+str(cell_row+2)+":S"+str(cell_row+2))  # A1 ~ F1까지 병합
        ws["A"+str(cell_row+2)] = "현장명 : " + site
        ws["A"+str(cell_row+2)].alignment = Alignment(horizontal='center')
        ws.merge_cells("A"+str(cell_row+3)+":D"+str(cell_row+3))  # A1 ~ F1까지 병합
        ws.merge_cells("E"+str(cell_row+3)+":S"+str(cell_row+3))  # A1 ~ F1까지 병합
        ws["A"+str(cell_row+3)] = "날짜 : " + start.split(" ")[0]

        ws.append(['NO', '공종', '성명', '공수', '# 주간작업', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
        for col in range(1,20):
            ws.cell(row=(cell_row+4), column=col).fill = gray_fill

        ws.merge_cells("E"+str(cell_row+4)+":S"+str(cell_row+4))  # 국적 병합
        ws["E"+str(cell_row+4)].alignment = Alignment(horizontal='center')

        ws.append(['', '', '', '', '구간', '층', '형태', '작업 단가별 분류', '작업내용', '공종', '공수', '작업명단', '', '', '', '', '', '', ''])
        ws["L"+str(cell_row+4)].alignment = Alignment(horizontal='center')
        ws.merge_cells("L"+str(cell_row+5)+":S"+str(cell_row+5)) 
        for col in range(5,20):
            ws.cell(row=(cell_row+5), column=col).fill = gray_fill
            
        full_number =0 

        for jj in range(cell_row+6,+cell_row+29):
            ws.merge_cells("L"+str(jj)+":S"+str(jj))

        for i in range(len(temp_parse)):
            parse_i = temp_parse[i].split("\n")
            print(parse_i)
            ##E~K 6부터 
            number = int(parse_i[1].split(":")[1].replace(" ","")) ## 공수 
            full_number = full_number+number ## 합계 

            work_parse = parse_i[2].split(":")[1].replace(" ","") ## 
            work1 = work_parse.split("_")[1] ##형태 
            floor = work_parse.split("_")[2] ## 층 
            floor2 = work_parse.split("_")[3] ## 구간
            work2 = work_parse.split("_")[4] ## 단가별 분류
            work3 = work_parse.split("_")[5] ## 작업내용 
            # for aa in ['E','F','G','H','I','J','K']:
            ws['E'+str(6+i+cell_row)] = floor2
            ws['F'+str(6+i+cell_row)] = floor
            ws['G'+str(6+i+cell_row)] = work1
            ws['H'+str(6+i+cell_row)] = work2
            ws['I'+str(6+i+cell_row)] = work3
            ws['J'+str(6+i+cell_row)] = '형틀'
            ws['K'+str(6+i+cell_row)] = number

            ## A~D 5부터 
            # for j in range(5,len(parse_i)):
                # print(j)
            index_of_table = None
            for i, item in enumerate(parse_i):
                # "테 이블 데이터:" 라는 문자열이 포함되어 있는지 확인
                if "테이블 데이터:" in item:
                    index_of_table = i+1
                    break
            for j in range(index_of_table,len(parse_i)):
                if parse_i[j]=="":
                    pass
                else:
                    today_worker = today_worker+[{parse_i[j].split(":")[1].split("|")[0].replace(" ",""):parse_i[j].split(":")[1].split("|")[1].replace(" ","")}]
                    print(today_worker)

        ws.merge_cells("E"+str(cell_row+29)+":J"+str(cell_row+29))
        ws["E"+str(cell_row+29)]='합계'
        ws["K"+str(cell_row+29)] = full_number
        
        ws.merge_cells("L"+str(cell_row+29)+":S"+str(cell_row+29))
        ws.merge_cells("E"+str(cell_row+30)+":S"+str(cell_row+30))
        ws["E"+str(cell_row+30)] = "# 연장작업"
        ws.cell(row=(cell_row+30), column=5).fill = gray_fill
        for col in range(5,20):
            ws.cell(row=(cell_row+31), column=col).fill = gray_fill
        ws["E"+str(cell_row+31)] = '구간'
        ws["F"+str(cell_row+31)] = '층'
        ws["G"+str(cell_row+31)] = '형태'
        ws["H"+str(cell_row+31)] = '작업 단가별 분류'
        ws["I"+str(cell_row+31)] = '작업내용'
        ws["J"+str(cell_row+31)] = '공종'
        ws["K"+str(cell_row+31)] = '공수'
        ws.merge_cells("L"+str(cell_row+31)+":S"+str(cell_row+31))
        ws["L"+str(cell_row+31)] = '작업 명단'
        for jj in range(cell_row+32,+cell_row+44):
            ws.merge_cells("L"+str(jj)+":S"+str(jj))

        for idx,k in enumerate(today_worker):
            ws['A'+str(5+idx+cell_row)] = idx+1
            ws['B'+str(5+idx+cell_row)] = '형틀'
            ws['C'+str(5+idx+cell_row)] = list(k.keys())[0]
            ws['D'+str(5+idx+cell_row)] = list(k.values())[0]
            last_k = idx
            last_idx = (5+idx+cell_row)

        for i,idx in enumerate(range(last_idx,44+cell_row)):
            ws['A'+str(idx)] = last_k+i
            ws['B'+str(idx)] = '형틀'

        for col in range(1,20):
            ws.cell(row=(cell_row+44), column=col).fill = gray_fill
            ws.cell(row=(cell_row+45), column=col).fill = gray_fill

        ws["A"+str(cell_row+44)] = '합계'
        ws.merge_cells("A"+str(cell_row+44)+":C"+str(cell_row+44))
        ws["D"+str(cell_row+44)] = full_number
        ws.merge_cells("E"+str(cell_row+44)+":J"+str(cell_row+44))
        ws["E"+str(cell_row+44)] = '합계'
        ws.merge_cells("L"+str(cell_row+44)+":S"+str(cell_row+44))
        ws["K"+str(cell_row+44)] = full_number
        ws["A"+str(cell_row+45)] = "# 특이사항"
        ws.merge_cells("A"+str(cell_row+45)+":S"+str(cell_row+45))
        for kk in range(46+cell_row,50+cell_row):
            ws.merge_cells('A'+str(kk)+':S'+str(kk))
            
        alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        # (2) 모든 사용된 셀 범위를 조회 (1~ws.max_row행, 1~8열) 후 테두리/정렬 적용
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=19):
            for cell in row:
                # 테두리
                cell.border = thin_border
                cell.alignment = alignment_center

    wb.save("./static/data/작업일지_"+date+".xlsx")

    try:
        return jsonify({'success': True, 'message': '이미지가 성공적으로 저장되었습니다.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=80,debug=True)
